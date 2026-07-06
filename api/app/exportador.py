"""Utilidades para exportar reportes a Excel (XLSX) y PDF.

- a_excel / a_pdf: un solo reporte (título + columnas + filas).
- a_excel_multi / a_pdf_multi: varios reportes en un solo archivo
  (una hoja por reporte en Excel; secciones apiladas en PDF).
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

CAFE = "6F4E37"  # color de marca para encabezados


def _estilo_tabla_pdf():
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + CAFE)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F0EB")]),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ])


def _volcar_hoja(ws, titulo, columnas, filas):
    ws.append([titulo])
    ws["A1"].font = Font(bold=True, size=14)

    ws.append(list(columnas))
    encabezado_fila = ws.max_row
    relleno = PatternFill(start_color=CAFE, end_color=CAFE, fill_type="solid")
    for celda in ws[encabezado_fila]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = relleno

    for fila in filas:
        ws.append(list(fila))

    for col in ws.columns:
        ancho = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = ancho + 4


def a_excel(titulo: str, columnas: list, filas: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"
    _volcar_hoja(ws, titulo, columnas, filas)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def a_excel_multi(hojas: list) -> bytes:
    """hojas: [{"nombre","titulo","columnas","filas"}] — una hoja por reporte."""
    wb = Workbook()
    wb.remove(wb.active)  # quitar la hoja por defecto
    for h in hojas:
        ws = wb.create_sheet(title=h["nombre"][:31])
        _volcar_hoja(ws, h["titulo"], h["columnas"], h["filas"])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def a_pdf(titulo: str, columnas: list, filas: list) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, title=titulo)
    estilos = getSampleStyleSheet()

    elementos = [Paragraph(titulo, estilos["Title"]), Spacer(1, 0.5 * cm)]
    datos = [list(columnas)] + [[str(c) for c in fila] for fila in filas]
    if not filas:
        datos.append(["Sin datos para el periodo seleccionado"])
    tabla = Table(datos, repeatRows=1)
    tabla.setStyle(_estilo_tabla_pdf())
    elementos.append(tabla)
    doc.build(elementos)
    return buffer.getvalue()


def a_pdf_multi(titulo: str, secciones: list) -> bytes:
    """secciones: [{"titulo","columnas","filas"}] — apiladas en un PDF."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, title=titulo)
    estilos = getSampleStyleSheet()

    elementos = [Paragraph(titulo, estilos["Title"]), Spacer(1, 0.6 * cm)]
    for s in secciones:
        elementos.append(Paragraph(s["titulo"], estilos["Heading2"]))
        elementos.append(Spacer(1, 0.2 * cm))
        datos = [list(s["columnas"])] + [[str(c) for c in fila] for fila in s["filas"]]
        if not s["filas"]:
            datos.append(["Sin datos"])
        tabla = Table(datos, repeatRows=1)
        tabla.setStyle(_estilo_tabla_pdf())
        elementos.append(tabla)
        elementos.append(Spacer(1, 0.6 * cm))
    doc.build(elementos)
    return buffer.getvalue()
